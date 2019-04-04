from tkinter import *
from tkinter import messagebox
from openpyxl import *
import time
import random

# =========================BACK-END================================


def error_handling(msg):
    if msg == "error!":
        messagebox.showerror("Error", "No Digits to compute!")
    elif msg == "division by zero!":
        messagebox.showerror("Error", "You can not divide a digit by zero!")


def btn_one():
    calc_entry.insert(END, "1")


def btn_two():
    calc_entry.insert(END, "2")


def btn_three():
    calc_entry.insert(END, "3")


def btn_four():
    calc_entry.insert(END, "4")


def btn_five():
    calc_entry.insert(END, "5")


def btn_six():
    calc_entry.insert(END, "6")


def btn_seven():
    calc_entry.insert(END, "7")


def btn_eight():
    calc_entry.insert(END, "8")


def btn_nine():
    calc_entry.insert(END, "9")


def btn_zero():
    calc_entry.insert(END, "0")


def btn_plus():
    calc_entry.insert(END, "+")


def btn_mult():
    calc_entry.insert(END, "*")


def btn_minus():
    calc_entry.insert(END, "-")


def btn_div():
    calc_entry.insert(END, "/")


def btn_clear():
    calc_entry.delete(0, END)


def equals_to():
    try:
        x = calc_entry_var1.get()
        y = eval(x)
        calc_entry.delete(0, END)
        calc_entry.insert(END, y)
    except ZeroDivisionError:
        error_handling("division by zero!")
    except SyntaxError:
        error_handling("error!")


def exit_program():
    root.destroy()


def reset_sales():
    try:
        check_btn_var1.set(0)
        check_btn_var2.set(0)
        check_btn_var3.set(0)
        check_btn_var4.set(0)
        check_btn_var5.set(0)
        check_btn_var6.set(0)
        check_btn_var7.set(0)
        check_btn_var8.set(0)
        check_btn_var9.set(0)
        check_btn_var10.set(0)
        check_btn_var11.set(0)
        check_btn_var12.set(0)

        entry_other_fruit.delete(0, END)

        global list_menu
        menu_opt_var1.set(list_menu[0])
        menu_opt_var2.set(list_menu[0])
        menu_opt_var3.set(list_menu[0])
        menu_opt_var4.set(list_menu[0])
        menu_opt_var5.set(list_menu[0])
        menu_opt_var6.set(list_menu[0])
        menu_opt_var7.set(list_menu[0])
        menu_opt_var8.set(list_menu[0])
        menu_opt_var9.set(list_menu[0])
        menu_opt_var10.set(list_menu[0])
        menu_opt_var11.set(list_menu[0])
        menu_opt_var12.set(list_menu[0])
        menu_opt_var13.set(list_menu[0])

        lf2_check_btn_var1.set(0)
        lf2_check_btn_var2.set(0)
        lf2_check_btn_var3.set(0)
        lf2_check_btn_var4.set(0)
        lf2_check_btn_var5.set(0)
        lf2_check_btn_var6.set(0)
        lf2_check_btn_var7.set(0)
        lf2_check_btn_var8.set(0)
        lf2_check_btn_var9.set(0)
        lf2_check_btn_var10.set(0)
        lf2_check_btn_var11.set(0)
        lf2_check_btn_var12.set(0)

        lf2_entry_other_fruitJ.delete(0, END)

        global lf2_list_menu
        lf2_menu_opt_var1.set(lf2_list_menu[0])
        lf2_menu_opt_var2.set(lf2_list_menu[0])
        lf2_menu_opt_var3.set(lf2_list_menu[0])
        lf2_menu_opt_var4.set(lf2_list_menu[0])
        lf2_menu_opt_var5.set(lf2_list_menu[0])
        lf2_menu_opt_var6.set(lf2_list_menu[0])
        lf2_menu_opt_var7.set(lf2_list_menu[0])
        lf2_menu_opt_var8.set(lf2_list_menu[0])
        lf2_menu_opt_var9.set(lf2_list_menu[0])
        lf2_menu_opt_var10.set(lf2_list_menu[0])
        lf2_menu_opt_var11.set(lf2_list_menu[0])
        lf2_menu_opt_var12.set(lf2_list_menu[0])
        lf2_menu_opt_var13.set(lf2_list_menu[0])

        entry_cost_fruits.delete(0, END)
        entry_cost_juice.delete(0, END)
        entry_total_cost.delete(0, END)
        tax_entry.delete(0, END)
        subtotal_entry.delete(0, END)
        total_entry.delete(0, END)
        receipt_field.delete("1.0", END)
        items_quantity_price.clear()
        entry_price_other_fruit.delete(0, END)
        entry_price_other_juice.delete(0, END)
        messagebox.showinfo("Next Customer!", "Serve the next customer please!")
    except NameError:
        messagebox.showerror("Error!", "Nothing to clear!")


def total_sales():
    # ================================ FRUITS ============================== #
    price_tag = {"mango": 20, "orange": 10, "apple": 25, "banana": 10, "watermelon": 20, "avocado": 10, "pineapple": 30,
                 "strawberry": 50, "guava": 25, "pawpaw": 50, "pears": 25, "pepino-melon": 30,
                 "other": entry_price_other_fruit_var.get()}

    global items_quantity_price
    items_quantity_price = []

    if check_btn_var1.get() == 1:
        mango_results = int(menu_opt_var1.get()) * price_tag["mango"]
        mango_tuple = ("mango", str(menu_opt_var1.get()), "20")
        items_quantity_price.append(mango_tuple)
    else:
        mango_results = 0

    if check_btn_var2.get() == 1:
        orange_results = int(menu_opt_var2.get()) * price_tag["orange"]
        orange_tuple = ("orange", str(menu_opt_var2.get()), "10")
        items_quantity_price.append(orange_tuple)
    else:
        orange_results = 0

    if check_btn_var3.get() == 1:
        apple_results = int(menu_opt_var3.get()) * price_tag["apple"]
        apple_tuple = ("apple", str(menu_opt_var3.get()), "25")
        items_quantity_price.append(apple_tuple)
    else:
        apple_results = 0

    if check_btn_var4.get() == 1:
        banana_results = int(menu_opt_var4.get()) * price_tag["banana"]
        banana_tuple = ("banana", str(menu_opt_var4.get()), "10")
        items_quantity_price.append(banana_tuple)
    else:
        banana_results = 0

    if check_btn_var5.get() == 1:
        watermelon_results = int(menu_opt_var5.get()) * price_tag["watermelon"]
        watermelon_tuple = ("watermelon", str(menu_opt_var5.get()), "20")
        items_quantity_price.append(watermelon_tuple)
    else:
        watermelon_results = 0

    if check_btn_var6.get() == 1:
        avocado_results = int(menu_opt_var6.get()) * price_tag["avocado"]
        avocado_tuple = ("avocado", str(menu_opt_var6.get()), "10")
        items_quantity_price.append(avocado_tuple)
    else:
        avocado_results = 0

    if check_btn_var7.get() == 1:
        pineapple_results = int(menu_opt_var7.get()) * price_tag["pineapple"]
        pineapple_tuple = ("pineapple", str(menu_opt_var7.get()), "30")
        items_quantity_price.append(pineapple_tuple)
    else:
        pineapple_results = 0

    if check_btn_var8.get() == 1:
        strawberry_results = int(menu_opt_var8.get()) * price_tag["strawberry"]
        strawberry_tuple = ("strawberry", str(menu_opt_var8.get()), "50")
        items_quantity_price.append(strawberry_tuple)
    else:
        strawberry_results = 0

    if check_btn_var9.get() == 1:
        guava_results = int(menu_opt_var9.get()) * price_tag["guava"]
        guava_tuple = ("guava", str(menu_opt_var9.get()), "25")
        items_quantity_price.append(guava_tuple)
    else:
        guava_results = 0

    if check_btn_var10.get() == 1:
        pawpaw_results = int(menu_opt_var10.get()) * price_tag["pawpaw"]
        pawpaw_tuple = ("pawpaw", str(menu_opt_var10.get()), "50")
        items_quantity_price.append(pawpaw_tuple)
    else:
        pawpaw_results = 0

    if check_btn_var11.get() == 1:
        pear_results = int(menu_opt_var11.get()) * price_tag["pears"]
        pears_tuple = ("pears", str(menu_opt_var11.get()), "25")
        items_quantity_price.append(pears_tuple)
    else:
        pear_results = 0

    if check_btn_var12.get() == 1:
        pepinomelon_results = int(menu_opt_var12.get()) * price_tag["pepino-melon"]
        pepino_melon_tuple = ("pepino-melon", str(menu_opt_var12.get()), "30")
        items_quantity_price.append(pepino_melon_tuple)
    else:
        pepinomelon_results = 0

    try:
        global other_item_results
        global total_cost_of_fruits
        if entry_other_fruit.get():
            other_item_results = int(menu_opt_var13.get()) * int(entry_price_other_fruit_var.get())
            other_tuple = (entry_other_fruit.get(), str(menu_opt_var13.get()), entry_price_other_fruit_var.get())
            items_quantity_price.append(other_tuple)
        else:
            other_item_results = 0

        total_cost_of_fruits = (mango_results + orange_results + apple_results + banana_results + watermelon_results +
                                avocado_results + pineapple_results + strawberry_results + guava_results +
                                pawpaw_results + pear_results + pepinomelon_results + other_item_results
                                )

        entry_cost_fruits.delete(0, END)
        entry_cost_fruits.insert(0, "KES. " + str(total_cost_of_fruits))
    except NameError:
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")

    # ================================ JUICE-FRUITS ============================== #

    juice_price_tag = {"virgin Punch": 150, "grape nectar": 150, "plum-ness": 100, "mango_cherry": 200,
                       "seasonal_fruit": 100, "apple_lemon": 200, "kale_pear": 100, "healthy_berry": 200,
                       "apple_milk": 250, "sour_plum": 100, "coconut_milk": 180, "citrus_boost": 200,
                       "other_juice": entry_price_other_juice_var.get()}

    if lf2_check_btn_var1.get() == 1:
        x1 = int(lf2_menu_opt_var1.get()) * juice_price_tag["virgin Punch"]
        virgin_punch_tuple = ("virgin Punch", str(lf2_menu_opt_var1.get()), "150")
        items_quantity_price.append(virgin_punch_tuple)
    else:
        x1 = 0

    if lf2_check_btn_var2.get() == 1:
        x2 = int(lf2_menu_opt_var2.get()) * juice_price_tag["grape nectar"]
        grape_nectar_tuple = ("grape nectar", str(lf2_menu_opt_var2.get()), "150")
        items_quantity_price.append(grape_nectar_tuple)
    else:
        x2 = 0

    if lf2_check_btn_var3.get() == 1:
        x3 = int(lf2_menu_opt_var3.get()) * juice_price_tag["plum-ness"]
        plum_ness_tuple = ("plum-ness", str(lf2_menu_opt_var3.get()), "100")
        items_quantity_price.append(plum_ness_tuple)
    else:
        x3 = 0

    if lf2_check_btn_var4.get() == 1:
        x4 = int(lf2_menu_opt_var4.get()) * juice_price_tag["mango_cherry"]
        mango_cherry_tuple = ("mango_cherry", str(lf2_menu_opt_var4.get()), "200")
        items_quantity_price.append(mango_cherry_tuple)
    else:
        x4 = 0

    if lf2_check_btn_var5.get() == 1:
        x5 = int(lf2_menu_opt_var5.get()) * juice_price_tag["seasonal_fruit"]
        seasonal_fruit_tuple = ("seasonal_fruit", str(lf2_menu_opt_var5.get()), "100")
        items_quantity_price.append(seasonal_fruit_tuple)
    else:
        x5 = 0

    if lf2_check_btn_var6.get() == 1:
        x6 = int(lf2_menu_opt_var6.get()) * juice_price_tag["apple_lemon"]
        apple_lemon_tuple = ("apple_lemon", str(lf2_menu_opt_var6.get()), "200")
        items_quantity_price.append(apple_lemon_tuple)
    else:
        x6 = 0

    if lf2_check_btn_var7.get() == 1:
        x7 = int(lf2_menu_opt_var7.get()) * juice_price_tag["kale_pear"]
        kale_pear_tuple = ("kale_pear", str(lf2_menu_opt_var7.get()), "100")
        items_quantity_price.append(kale_pear_tuple)
    else:
        x7 = 0

    if lf2_check_btn_var8.get() == 1:
        x8 = int(lf2_menu_opt_var8.get()) * juice_price_tag["healthy_berry"]
        healthy_berry_tuple = ("healthy_berry", str(lf2_menu_opt_var8.get()), "200")
        items_quantity_price.append(healthy_berry_tuple)
    else:
        x8 = 0

    if lf2_check_btn_var9.get() == 1:
        x9 = int(lf2_menu_opt_var9.get()) * juice_price_tag["apple_milk"]
        apple_milk_tuple = ("apple_milk", str(lf2_menu_opt_var9.get()), "250")
        items_quantity_price.append(apple_milk_tuple)
    else:
        x9 = 0

    if lf2_check_btn_var10.get() == 1:
        x10 = int(lf2_menu_opt_var10.get()) * juice_price_tag["sour_plum"]
        sour_plum_tuple = ("sour_plum", str(lf2_menu_opt_var10.get()), "100")
        items_quantity_price.append(sour_plum_tuple)
    else:
        x10 = 0

    if lf2_check_btn_var11.get() == 1:
        x11 = int(lf2_menu_opt_var11.get()) * juice_price_tag["coconut_milk"]
        coconut_milk_tuple = ("coconut_milk", str(lf2_menu_opt_var11.get()), "180")
        items_quantity_price.append(coconut_milk_tuple)
    else:
        x11 = 0

    if lf2_check_btn_var12.get() == 1:
        x12 = int(lf2_menu_opt_var12.get()) * juice_price_tag["citrus_boost"]
        citrus_boost_tuple = ("citrus_boost", str(lf2_menu_opt_var12.get()), "200")
        items_quantity_price.append(citrus_boost_tuple)
    else:
        x12 = 0

    try:
        global x13
        if lf2_entry_other_fruitJ.get():
            x13 = int(lf2_menu_opt_var13.get()) * int(entry_price_other_juice_var.get())
            other_juice_tuple = (lf2_entry_other_fruitJ.get(), str(lf2_menu_opt_var13.get()),
                                 entry_price_other_juice_var.get())
            items_quantity_price.append(other_juice_tuple)
        else:
            x13 = 0

        global total_cost_of_juice
        global tax_on_all_items
        total_cost_of_juice = (x1 + x2 + x3 + x4 + x5 + x6 + x7 + x8 + x9 + x10 + x11 + x12 + x13)

        entry_cost_juice.delete(0, END)
        entry_cost_juice.insert(0, "KES. " + str(total_cost_of_juice))

        entry_total_cost.delete(0, END)
        entry_total_cost.insert(0, "KES. " + str(total_cost_of_juice + total_cost_of_fruits))

        tax_on_all_items = 0.16 * (total_cost_of_fruits + total_cost_of_juice)

        tax_entry.delete(0, END)
        tax_entry.insert(0, "KES. " + str(tax_on_all_items))

        subtotal_entry.delete(0, END)
        subtotal_entry.insert(0, "KES. " + str(total_cost_of_juice + total_cost_of_fruits))

        total_entry.delete(0, END)
        total_entry.insert(0, "KES. " + str(total_cost_of_juice + total_cost_of_fruits + tax_on_all_items))
    except NameError:
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")


def print_receipt():
    try:
        receipt_field.delete("1.0", END)
        receipt_field.insert(END, "  FRUITS & JUICE SHOP." + "\n")
        receipt_field.insert(END, " " + "\n")

        x = 0
        while True:
            receipt_field.insert(END, items_quantity_price[x][0] + " :   " + items_quantity_price[x][1] + " each @ " +
                                 items_quantity_price[x][2] + "\n")
            x += 1

            if x < len(items_quantity_price):
                continue
            else:
                break

        receipt_field.insert(END, " " + "\n")
        receipt_field.insert(END, "Cost of fruits: " + entry_cost_fruits_var.get() + "\n")
        receipt_field.insert(END, "Cost of juice: " + entry_cost_juice_var.get() + "\n")
        receipt_field.insert(END, "Total cost:  " + entry_total_cost_var.get() + "\n")
        receipt_field.insert(END, " " + "\n")
        receipt_field.insert(END, " ***** WELCOME AGAIN *****" + "\n")
    except IndexError:
        messagebox.showerror("Error!", "select items first! Thank you!")
    except NameError:
        messagebox.showerror("Error!", "select items first! Thank you!")


def book_keeping():
    # LOAD THE WORKBOOK
    try:
        global book
        destination_file = "POS_fruit_juice.xlsx"
        book = load_workbook(filename=destination_file)

        # GRABBING THE ACTIVE SHEET.
        sheet = book.active

        # USEFUL VARIABLES.
        row = 2
        column = 1
        global today
        today = time.strftime("%x")
        id_ = random.randint(1, 1000000)

        if items_quantity_price:
            while True:
                if sheet.cell(row=row, column=column).value:
                    row += 1
                    continue
                else:
                    sheet.cell(row=row, column=column).value = today
                    column += 1
                    sheet.cell(row=row, column=column).value = id_
                    column += 1

                    # adding elements to the item cell
                    x = len(items_quantity_price)
                    sheet.cell(row=row, column=column).value = x
                    column += 1

                    sheet.cell(row=row, column=column).value = total_cost_of_juice + total_cost_of_fruits
                    column += 1
                    sheet.cell(row=row, column=column).value = tax_on_all_items

                    if column == 5:
                        break

            book.save(filename=destination_file)
            messagebox.showinfo("Book-keeping", "DONE!")
        else:
            messagebox.showerror("Error!", "Nothing to add!")
    except NameError:
        messagebox.showerror("Error!", "Nothing to add!")


def today_report():
    destination_file2 = "POS_fruit_juice.xlsx"
    work_book = load_workbook(filename=destination_file2)

    # GRABBING THE ACTIVE SHEET.
    sheet2 = work_book.active

    row = 2
    column = 1
    max_row = sheet2.max_row
    total_amount = []
    total_vat = []
    today2 = time.strftime("%x")

    x = 0
    while x < max_row:
        if sheet2.cell(row=row, column=column).value == today2:
            total_amount.append(sheet2.cell(row=row, column=4).value)
            total_vat.append(sheet2.cell(row=row, column=5).value)
            row += 1
            x += 1
            continue
        else:
            row += 1
            x += 1
            continue

    answer1 = 0
    for integer in total_amount:
        answer1 += integer

    answer2 = 0
    for float_point in total_vat:
        answer2 += float_point

    receipt_field.delete("1.0", END)
    receipt_field.insert(END, "FRUITS & JUICE SHOP.          " + today2 + "." + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "         ***TODAY\'S REPORT.***" + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Total Amount generated today: " + "KES. " + str(answer1) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Total V.A.T generated today: " + "KES. " + str(answer2) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space
    receipt_field.insert(END, "Gross income + V.A.T: " + "KES. " + str(answer1 + answer2) + "\n")
    receipt_field.insert(END, "  " + "\n")  # blank space

# ======================BACK-END=============================


root = Tk()
# ==========================SETTINGS======================
root.resizable(width=FALSE, height=FALSE)
root.geometry("1200x650")
root.title("BILLING SYSTEM")
color = "Light Gray"
color2 = "Gray"
root.configure(bg=color)

# ==========================FRAMES=========================

# heading
top_frame = Frame(root, bg=color, width=1200, height=50)
top_frame.pack(side=TOP)

# footer for buttons
bottom_frame = Frame(root, bg=color, width=1200, height=150)
bottom_frame.pack(side=BOTTOM)

# fruits
left_frame = Frame(root, bg=color, width=300, height=500)
left_frame.pack(side=LEFT)

# fruit juice
left_frame2 = Frame(root, bg=color, width=300, height=500)
left_frame2.pack(side=LEFT, padx=5)

# billing results
right_frame = Frame(root, bg=color, width=800, height=150)
right_frame.place(x=610, y=60)

# receipt
right_frame2 = Frame(root, bg=color, width=800, height=280)
right_frame2.place(x=610, y=210)

# ============================TITLE on top FRAME=======================
title_label = Label(top_frame, text="Fruits & Juice shop.", font=("Aisha Latin Semibold", 20, "bold"), bg=color,
                    pady=10)
title_label.pack(side=LEFT)

# ==============================LABELS & ENTRIES on right frame========================
label_cost_fruits = Label(right_frame, text="Cost of fruits: ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
label_cost_fruits.place(x=10, y=5)

entry_cost_fruits_var = StringVar()
entry_cost_fruits = Entry(right_frame, textvariable=entry_cost_fruits_var, width=10)
entry_cost_fruits.place(x=160, y=5)


label_cost_juice = Label(right_frame, text="Cost of Juice : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
label_cost_juice.place(x=10, y=50)

entry_cost_juice_var = StringVar()
entry_cost_juice = Entry(right_frame, textvariable=entry_cost_juice_var, width=10)
entry_cost_juice.place(x=160, y=50)


label_total_cost = Label(right_frame, text="Total Cost     : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
label_total_cost.place(x=10, y=95)

entry_total_cost_var = StringVar()
entry_total_cost = Entry(right_frame, textvariable=entry_total_cost_var, width=10)
entry_total_cost.place(x=160, y=95)


tax_label = Label(right_frame, text="Tax :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
tax_label.place(x=350, y=5)

tax_var = StringVar()
tax_entry = Entry(right_frame, textvariable=tax_var, width=10)
tax_entry.place(x=410, y=5)


subtotal = Label(right_frame, text="Sub Total:", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
subtotal.place(x=350, y=50)

subtotal_var = StringVar()
subtotal_entry = Entry(right_frame, textvariable=subtotal_var, width=10)
subtotal_entry.place(x=460, y=50)


total = Label(right_frame, text="Total:", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
total.place(x=350, y=95)

total_var = StringVar()
total_entry = Entry(right_frame, textvariable=total_var, width=10)
total_entry.place(x=420, y=95)

# ==============================Text field on right_frame2========================
receipt_field = Text(right_frame2, width=58, height=14)
receipt_field.pack(side=TOP)

# ==============================Check-boxes, label, & an entry on the left_frame=====================

check_btn_var1 = IntVar()
check_btn1 = Checkbutton(left_frame, text="Mango", variable=check_btn_var1)
check_btn1.place(x=3, y=5)

check_btn_var2 = IntVar()
check_btn2 = Checkbutton(left_frame, text="Orange", variable=check_btn_var2)
check_btn2.place(x=3, y=35)

check_btn_var3 = IntVar()
check_btn3 = Checkbutton(left_frame, text="Apple", variable=check_btn_var3)
check_btn3.place(x=3, y=65)

check_btn_var4 = IntVar()
check_btn4 = Checkbutton(left_frame, text="Banana", variable=check_btn_var4)
check_btn4.place(x=3, y=95)

check_btn_var5 = IntVar()
check_btn5 = Checkbutton(left_frame, text="Watermelon", variable=check_btn_var5)
check_btn5.place(x=3, y=125)

check_btn_var6 = IntVar()
check_btn6 = Checkbutton(left_frame, text="Avocado", variable=check_btn_var6)
check_btn6.place(x=3, y=155)

check_btn_var7 = IntVar()
check_btn7 = Checkbutton(left_frame, text="Pineapple", variable=check_btn_var7)
check_btn7.place(x=3, y=185)

check_btn_var8 = IntVar()
check_btn8 = Checkbutton(left_frame, text="Strawberry", variable=check_btn_var8)
check_btn8.place(x=3, y=215)

check_btn_var9 = IntVar()
check_btn9 = Checkbutton(left_frame, text="Guava", variable=check_btn_var9)
check_btn9.place(x=3, y=245)

check_btn_var10 = IntVar()
check_btn10 = Checkbutton(left_frame, text="Paw-paw", variable=check_btn_var10)
check_btn10.place(x=3, y=275)

check_btn_var11 = IntVar()
check_btn11 = Checkbutton(left_frame, text="Pears", variable=check_btn_var11)
check_btn11.place(x=3, y=305)

check_btn_var12 = IntVar()
check_btn12 = Checkbutton(left_frame, text="Pepino Melon", variable=check_btn_var12)
check_btn12.place(x=3, y=335)

label_other_fruit = Label(left_frame, text="Other :")
label_other_fruit.place(x=3, y=390)

entry_other_fruit = Entry(left_frame, width=10)
entry_other_fruit.place(x=100, y=390)


# ======================menu-options on the left frame=============================
list_menu = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]

# mango
menu_opt_var1 = StringVar()
menu_opt_var1.set(list_menu[0])
menu_opt1 = OptionMenu(left_frame, menu_opt_var1, *list_menu)
menu_opt1.configure(bg=color, font=("arial", 7))
menu_opt1.place(x=200, y=5)

menu_opt_var2 = StringVar()
menu_opt_var2.set(list_menu[0])
menu_opt2 = OptionMenu(left_frame, menu_opt_var2, *list_menu)
menu_opt2.configure(bg=color, font=("arial", 7))
menu_opt2.place(x=200, y=35)

menu_opt_var3 = StringVar()
menu_opt_var3.set(list_menu[0])
menu_opt3 = OptionMenu(left_frame, menu_opt_var3, *list_menu)
menu_opt3.configure(bg=color, font=("arial", 7))
menu_opt3.place(x=200, y=65)

menu_opt_var4 = StringVar()
menu_opt_var4.set(list_menu[0])
menu_opt4 = OptionMenu(left_frame, menu_opt_var4, *list_menu)
menu_opt4.configure(bg=color, font=("arial", 7))
menu_opt4.place(x=200, y=95)

menu_opt_var5 = StringVar()
menu_opt_var5.set(list_menu[0])
menu_opt5 = OptionMenu(left_frame, menu_opt_var5, *list_menu)
menu_opt5.configure(bg=color, font=("arial", 7))
menu_opt5.place(x=200, y=125)

menu_opt_var6 = StringVar()
menu_opt_var6.set(list_menu[0])
menu_opt6 = OptionMenu(left_frame, menu_opt_var6, *list_menu)
menu_opt6.configure(bg=color, font=("arial", 7))
menu_opt6.place(x=200, y=155)

menu_opt_var7 = StringVar()
menu_opt_var7.set(list_menu[0])
menu_opt7 = OptionMenu(left_frame, menu_opt_var7, *list_menu)
menu_opt7.configure(bg=color, font=("arial", 7))
menu_opt7.place(x=200, y=185)

menu_opt_var8 = StringVar()
menu_opt_var8.set(list_menu[0])
menu_opt8 = OptionMenu(left_frame, menu_opt_var8, *list_menu)
menu_opt8.configure(bg=color, font=("arial", 7))
menu_opt8.place(x=200, y=215)

menu_opt_var9 = StringVar()
menu_opt_var9.set(list_menu[0])
menu_opt9 = OptionMenu(left_frame, menu_opt_var9, *list_menu)
menu_opt9.configure(bg=color, font=("arial", 7))
menu_opt9.place(x=200, y=245)

menu_opt_var10 = StringVar()
menu_opt_var10.set(list_menu[0])
menu_opt10 = OptionMenu(left_frame, menu_opt_var10, *list_menu)
menu_opt10.configure(bg=color, font=("arial", 7))
menu_opt10.place(x=200, y=275)

menu_opt_var11 = StringVar()
menu_opt_var11.set(list_menu[0])
menu_opt11 = OptionMenu(left_frame, menu_opt_var11, *list_menu)
menu_opt11.configure(bg=color, font=("arial", 7))
menu_opt11.place(x=200, y=305)

menu_opt_var12 = StringVar()
menu_opt_var12.set(list_menu[0])
menu_opt12 = OptionMenu(left_frame, menu_opt_var12, *list_menu)
menu_opt12.configure(bg=color, font=("arial", 7))
menu_opt12.place(x=200, y=335)

menu_opt_var13 = StringVar()
menu_opt_var13.set(list_menu[0])
menu_opt13 = OptionMenu(left_frame, menu_opt_var13, *list_menu)
menu_opt13.configure(bg=color, font=("arial", 7))
menu_opt13.place(x=240, y=387)


# ==============================Check-boxes, label, & an entry on the left_frame2=====================

lf2_check_btn_var1 = IntVar()
lf2_check_btn1 = Checkbutton(left_frame2, text="Virgin Punch", variable=lf2_check_btn_var1)
lf2_check_btn1.place(x=3, y=5)

lf2_check_btn_var2 = IntVar()
lf2_check_btn2 = Checkbutton(left_frame2, text="Grape Nectar", variable=lf2_check_btn_var2)
lf2_check_btn2.place(x=3, y=35)

lf2_check_btn_var3 = IntVar()
lf2_check_btn3 = Checkbutton(left_frame2, text="Plum-ness", variable=lf2_check_btn_var3)
lf2_check_btn3.place(x=3, y=65)

lf2_check_btn_var4 = IntVar()
lf2_check_btn4 = Checkbutton(left_frame2, text="Mango + Cherry", variable=lf2_check_btn_var4)
lf2_check_btn4.place(x=3, y=95)

lf2_check_btn_var5 = IntVar()
lf2_check_btn5 = Checkbutton(left_frame2, text="Seasonal Fruit", variable=lf2_check_btn_var5)
lf2_check_btn5.place(x=3, y=125)

lf2_check_btn_var6 = IntVar()
lf2_check_btn6 = Checkbutton(left_frame2, text="Apple + Lemon", variable=lf2_check_btn_var6)
lf2_check_btn6.place(x=3, y=155)

lf2_check_btn_var7 = IntVar()
lf2_check_btn7 = Checkbutton(left_frame2, text="Kale + Pear", variable=lf2_check_btn_var7)
lf2_check_btn7.place(x=3, y=185)

lf2_check_btn_var8 = IntVar()
lf2_check_btn8 = Checkbutton(left_frame2, text="Healthy Berry", variable=lf2_check_btn_var8)
lf2_check_btn8.place(x=3, y=215)

lf2_check_btn_var9 = IntVar()
lf2_check_btn9 = Checkbutton(left_frame2, text="Apple + Milk", variable=lf2_check_btn_var9)
lf2_check_btn9.place(x=3, y=245)

lf2_check_btn_var10 = IntVar()
lf2_check_btn10 = Checkbutton(left_frame2, text="Sour Plum", variable=lf2_check_btn_var10)
lf2_check_btn10.place(x=3, y=275)

lf2_check_btn_var11 = IntVar()
lf2_check_btn11 = Checkbutton(left_frame2, text="Coconut milk", variable=lf2_check_btn_var11)
lf2_check_btn11.place(x=3, y=305)

lf2_check_btn_var12 = IntVar()
lf2_check_btn12 = Checkbutton(left_frame2, text="Citrus Boost", variable=lf2_check_btn_var12)
lf2_check_btn12.place(x=3, y=335)

lf2_label_other_fruitJ = Label(left_frame2, text="Other :")
lf2_label_other_fruitJ.place(x=3, y=390)

lf2_entry_other_fruitJ = Entry(left_frame2, width=10)
lf2_entry_other_fruitJ.place(x=100, y=390)


# ======================menu-options on the left frame2=============================

lf2_list_menu = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]

lf2_menu_opt_var1 = StringVar()
lf2_menu_opt_var1.set(list_menu[0])
lf2_menu_opt1 = OptionMenu(left_frame2, lf2_menu_opt_var1, *lf2_list_menu)
lf2_menu_opt1.configure(bg=color, font=("arial", 7))
lf2_menu_opt1.place(x=200, y=5)

lf2_menu_opt_var2 = StringVar()
lf2_menu_opt_var2.set(list_menu[0])
lf2_menu_opt2 = OptionMenu(left_frame2, lf2_menu_opt_var2, *lf2_list_menu)
lf2_menu_opt2.configure(bg=color, font=("arial", 7))
lf2_menu_opt2.place(x=200, y=35)

lf2_menu_opt_var3 = StringVar()
lf2_menu_opt_var3.set(list_menu[0])
lf2_menu_opt3 = OptionMenu(left_frame2, lf2_menu_opt_var3, *lf2_list_menu)
lf2_menu_opt3.configure(bg=color, font=("arial", 7))
lf2_menu_opt3.place(x=200, y=65)

lf2_menu_opt_var4 = StringVar()
lf2_menu_opt_var4.set(list_menu[0])
lf2_menu_opt4 = OptionMenu(left_frame2, lf2_menu_opt_var4, *lf2_list_menu)
lf2_menu_opt4.configure(bg=color, font=("arial", 7))
lf2_menu_opt4.place(x=200, y=95)

lf2_menu_opt_var5 = StringVar()
lf2_menu_opt_var5.set(list_menu[0])
lf2_menu_opt5 = OptionMenu(left_frame2, lf2_menu_opt_var5, *lf2_list_menu)
lf2_menu_opt5.configure(bg=color, font=("arial", 7))
lf2_menu_opt5.place(x=200, y=125)

lf2_menu_opt_var6 = StringVar()
lf2_menu_opt_var6.set(list_menu[0])
lf2_menu_opt6 = OptionMenu(left_frame2, lf2_menu_opt_var6, *lf2_list_menu)
lf2_menu_opt6.configure(bg=color, font=("arial", 7))
lf2_menu_opt6.place(x=200, y=155)

lf2_menu_opt_var7 = StringVar()
lf2_menu_opt_var7.set(list_menu[0])
lf2_menu_opt7 = OptionMenu(left_frame2, lf2_menu_opt_var7, *lf2_list_menu)
lf2_menu_opt7.configure(bg=color, font=("arial", 7))
lf2_menu_opt7.place(x=200, y=185)

lf2_menu_opt_var8 = StringVar()
lf2_menu_opt_var8.set(list_menu[0])
lf2_menu_opt8 = OptionMenu(left_frame2, lf2_menu_opt_var8, *lf2_list_menu)
lf2_menu_opt8.configure(bg=color, font=("arial", 7))
lf2_menu_opt8.place(x=200, y=215)

lf2_menu_opt_var9 = StringVar()
lf2_menu_opt_var9.set(list_menu[0])
lf2_menu_opt9 = OptionMenu(left_frame2, lf2_menu_opt_var9, *lf2_list_menu)
lf2_menu_opt9.configure(bg=color, font=("arial", 7))
lf2_menu_opt9.place(x=200, y=245)

lf2_menu_opt_var10 = StringVar()
lf2_menu_opt_var10.set(list_menu[0])
lf2_menu_opt10 = OptionMenu(left_frame2, lf2_menu_opt_var10, *lf2_list_menu)
lf2_menu_opt10.configure(bg=color, font=("arial", 7))
lf2_menu_opt10.place(x=200, y=275)

lf2_menu_opt_var11 = StringVar()
lf2_menu_opt_var11.set(list_menu[0])
lf2_menu_opt11 = OptionMenu(left_frame2, lf2_menu_opt_var11, *lf2_list_menu)
lf2_menu_opt11.configure(bg=color, font=("arial", 7))
lf2_menu_opt11.place(x=200, y=305)

lf2_menu_opt_var12 = StringVar()
lf2_menu_opt_var12.set(list_menu[0])
lf2_menu_opt12 = OptionMenu(left_frame2, lf2_menu_opt_var12, *lf2_list_menu)
lf2_menu_opt12.configure(bg=color, font=("arial", 7))
lf2_menu_opt12.place(x=200, y=335)

lf2_menu_opt_var13 = StringVar()
lf2_menu_opt_var13.set(list_menu[0])
lf2_menu_opt13 = OptionMenu(left_frame2, lf2_menu_opt_var13, *lf2_list_menu)
lf2_menu_opt13.configure(bg=color, font=("arial", 7))
lf2_menu_opt13.place(x=240, y=387)


# ================================BUTTONS FOR COMPUTATION and entries for pricing other items.=====================
entry_price_other_fruit_var = StringVar()
entry_price_other_fruit = Entry(bottom_frame, textvariable=entry_price_other_fruit_var, width=10)
entry_price_other_fruit.place(x=100, y=5)

entry_price_other_juice_var = StringVar()
entry_price_other_juice = Entry(bottom_frame, textvariable=entry_price_other_juice_var, width=10)
entry_price_other_juice.place(x=405, y=5)

total_btn_btm_frame = Button(bottom_frame, text="Total", width=10, bg=color2, command=lambda: total_sales())
total_btn_btm_frame.place(x=5, y=60)

receipt_btn_btm_frame = Button(bottom_frame, text="Receipt", width=10, bg=color2, command=lambda: print_receipt())
receipt_btn_btm_frame.place(x=150, y=60)

reset_btn_btm_frame = Button(bottom_frame, text="Reset", width=10, bg=color2, command=lambda: reset_sales())
reset_btn_btm_frame.place(x=295, y=60)

book_k_btn_btm_frame = Button(bottom_frame, text="Book keep", width=10, bg=color2, command=lambda: book_keeping())
book_k_btn_btm_frame.place(x=5, y=115)

report_btn_btm_frame = Button(bottom_frame, text="Today\'s report", width=10, bg=color2, command=lambda: today_report())
report_btn_btm_frame.place(x=150, y=115)

exit_btn_btm_frame = Button(bottom_frame, text="EXIT", width=10, bg=color2, command=lambda: exit_program())
exit_btn_btm_frame.place(x=295, y=115)

# ==================================THE CALCULATOR===================================

calc_entry_var1 = StringVar()
calc_entry = Entry(bottom_frame, textvariable=calc_entry_var1, width=44)
calc_entry.place(x=600, y=5)

calc_btn1 = Button(bottom_frame, text="1", width=10, bg=color2, font=("arial", 7), command=lambda: btn_one())
calc_btn1.place(x=600, y=40)

calc_btn2 = Button(bottom_frame, text="2", width=10, bg=color2, font=("arial", 7), command=lambda: btn_two())
calc_btn2.place(x=700, y=40)

calc_btn3 = Button(bottom_frame, text="3", width=10, bg=color2, font=("arial", 7), command=lambda: btn_three())
calc_btn3.place(x=800, y=40)

calc_btn4 = Button(bottom_frame, text="4", width=10, bg=color2, font=("arial", 7), command=lambda: btn_four())
calc_btn4.place(x=900, y=40)

calc_btn5 = Button(bottom_frame, text="5", width=10, bg=color2, font=("arial", 7), command=lambda: btn_five())
calc_btn5.place(x=1000, y=40)

calc_btn6 = Button(bottom_frame, text="6", width=10, bg=color2, font=("arial", 7), command=lambda: btn_six())
calc_btn6.place(x=600, y=80)

calc_btn7 = Button(bottom_frame, text="7", width=10, bg=color2, font=("arial", 7), command=lambda: btn_seven())
calc_btn7.place(x=700, y=80)

calc_btn8 = Button(bottom_frame, text="8", width=10, bg=color2, font=("arial", 7), command=lambda: btn_eight())
calc_btn8.place(x=800, y=80)

calc_btn9 = Button(bottom_frame, text="9", width=10, bg=color2, font=("arial", 7), command=lambda: btn_nine())
calc_btn9.place(x=900, y=80)

calc_btn0 = Button(bottom_frame, text="0", width=10, bg=color2, font=("arial", 7), command=lambda: btn_zero())
calc_btn0.place(x=1000, y=80)

calc_btn_plus = Button(bottom_frame, text="+", width=10, bg=color2, font=("arial", 7), command=lambda: btn_plus())
calc_btn_plus.place(x=600, y=120)

calc_btn_mult = Button(bottom_frame, text="*", width=10, bg=color2, font=("arial", 7), command=lambda: btn_mult())
calc_btn_mult.place(x=700, y=120)

calc_btn_minus = Button(bottom_frame, text="-", width=10, bg=color2, font=("arial", 7), command=lambda: btn_minus())
calc_btn_minus.place(x=800, y=120)

calc_btn_div = Button(bottom_frame, text="/", width=10, bg=color2, font=("arial", 7), command=lambda: btn_div())
calc_btn_div.place(x=900, y=120)

calc_btn_clear = Button(bottom_frame, text="C", width=10, bg=color2, font=("arial", 7), command=lambda: btn_clear())
calc_btn_clear.place(x=1000, y=120)

equals_to_button = Button(bottom_frame, text="=", width=5, height=3, bg=color2, font=("arial", 15),
                          command=lambda: equals_to())
equals_to_button.place(x=1100, y=45)

root.mainloop()
